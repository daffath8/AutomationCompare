import os
import json
import pandas as pd
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


json_folder = "/Users/daffath/Documents/AutomationCompare/Data/Data_json"  # folder hasil ekstrak zip
csv_path = "/Users/daffath/Documents/AutomationCompare/Data/arti.csv"
output_file = "/Users/daffath/Documents/AutomationCompare/Result/arti_comparison.xlsx"


os.makedirs(os.path.dirname(output_file), exist_ok=True)

# Fungsi normalisasi teks
def normalize(text):
    if not isinstance(text, str):
        return ""
    return unicodedata.normalize("NFKC", text.strip())

# Load CSV
csv_data = pd.read_csv(csv_path)
total_csv_rows = len(csv_data)

results = []
row_index = 0


for i in range(1, 115):
    filename = f"{i}.json"
    json_path = os.path.join(json_folder, filename)

    if not os.path.exists(json_path):
        print(f"⚠️ File JSON tidak ditemukan: {filename}")
        continue

    # Baca file JSON
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    ayat_list = data.get("ayat", [])

    # Loop setiap ayat di dalam file JSON
    for ayat in ayat_list:
        teks_json = normalize(ayat.get("teksIndonesia", ""))

        if row_index >= total_csv_rows:
            print(f"❗ Baris CSV tidak cukup untuk ayat ke-{row_index+1}")
            break

        teks_csv = normalize(csv_data.iloc[row_index]["teks_terjemah_2019"])
        status = "MATCH" if teks_json == teks_csv else "NOT_MATCH"

        results.append({
            "teksIndonesia (JSON)": teks_json,
            "teks_terjemah_2019 (CSV)": teks_csv,
            "validate": status
        })

        row_index += 1

# Simpan ke Excel
df_result = pd.DataFrame(results)
df_result.to_excel(output_file, index=False)

# === Tambahkan Highlight Warna ===
wb = load_workbook(output_file)
ws = wb.active

# Penentuan Warna untuk membedakan validasi
fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Hijau untuk MATCH
fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Merah untuk NOT_MATCH

# Cari kolom "validate"
validate_col_idx = None
for idx, col in enumerate(ws[1], start=1):
    if col.value == "validate":
        validate_col_idx = idx
        break

if validate_col_idx:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=validate_col_idx)
        if cell.value == "MATCH":
            cell.fill = fill_green
        elif cell.value == "NOT_MATCH":
            cell.fill = fill_red

wb.save(output_file)

print(f"\n✅ Hasil perbandingan dengan highlight disimpan di:\n{output_file}")
